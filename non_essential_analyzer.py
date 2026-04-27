"""
TAXY-Z — Non-Essential Spending Analyzer
==========================================
Identifies non-essential expenses across multiple years and generates
a savings-opportunity analysis Excel report.

Usage:
    from non_essential_analyzer import build_non_essential_report
    build_non_essential_report(all_txns, years=[2021,2022,2023], output_path="analysis.xlsx")
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from typing import List, Dict

# Color map per group
GROUP_STYLES = {
    "FOOD & DINING":        {"bg": "FCE4D6", "header": "C00000"},
    "SHOPPING":             {"bg": "FFF2CC", "header": "B7472A"},
    "SUBSCRIPTIONS":        {"bg": "DAE3F3", "header": "2E75B6"},
    "TRAVEL":               {"bg": "E2EFDA", "header": "375623"},
    "ENTERTAINMENT":        {"bg": "EAD1DC", "header": "833C50"},
    "PERSONAL CARE & MISC": {"bg": "F2F2F2", "header": "444444"},
}

# Mapping of category strings to groups
CATEGORY_GROUPS = {
    "FOOD & DINING": [
        "NON-ESSENTIAL – Food Delivery (DoorDash)",
        "NON-ESSENTIAL – Food Delivery (Groceries via DD)",
        "NON-ESSENTIAL – Restaurants & Dining Out",
    ],
    "SHOPPING": [
        "NON-ESSENTIAL – Shopping (Amazon)",
        "NON-ESSENTIAL – Shopping (Clothing)",
        "NON-ESSENTIAL – Shopping (Beauty)",
        "NON-ESSENTIAL – Shopping (Jewelry)",
        "NON-ESSENTIAL – Shopping (Electronics)",
        "NON-ESSENTIAL – Shopping (eBay)",
        "NON-ESSENTIAL – Shopping (Sports Equipment)",
        "NON-ESSENTIAL – Shopping (Home/Furniture)",
        "NON-ESSENTIAL – Shopping (Walmart)",
        "NON-ESSENTIAL – Shopping (Arts & Crafts)",
        "NON-ESSENTIAL – Shopping (Gifts/Flowers)",
        "NON-ESSENTIAL – Shopping (Hardware)",
        "NON-ESSENTIAL – Shopping (Storage Unit)",
        "NON-ESSENTIAL – Shopping (Misc)",
    ],
    "SUBSCRIPTIONS": [
        "NON-ESSENTIAL – Subscriptions (Amazon Prime)",
        "NON-ESSENTIAL – Subscriptions (Streaming)",
        "NON-ESSENTIAL – Subscriptions (Gaming)",
        "NON-ESSENTIAL – Subscriptions (Kids Sports)",
        "NON-ESSENTIAL – Subscriptions (Apple)",
        "NON-ESSENTIAL – Subscriptions (Google)",
        "NON-ESSENTIAL – Subscriptions (Microsoft)",
        "NON-ESSENTIAL – Subscriptions (Title Lock)",
        "NON-ESSENTIAL – Subscriptions (Hair Care)",
        "NON-ESSENTIAL – Subscriptions (Kindle)",
        "NON-ESSENTIAL – Subscriptions (Web Hosting)",
        "NON-ESSENTIAL – Subscriptions (AI/Tech)",
        "NON-ESSENTIAL – Subscriptions (Other)",
    ],
    "TRAVEL": [
        "NON-ESSENTIAL – Travel (Flights)",
        "NON-ESSENTIAL – Travel (Booking)",
        "NON-ESSENTIAL – Travel (Airbnb)",
        "NON-ESSENTIAL – Travel (Hotel)",
        "NON-ESSENTIAL – Travel (Catalina)",
        "NON-ESSENTIAL – Travel (Parking)",
        "NON-ESSENTIAL – Travel (Insurance)",
    ],
    "ENTERTAINMENT": [
        "NON-ESSENTIAL – Entertainment (Events)",
        "NON-ESSENTIAL – Entertainment (Activities)",
    ],
    "PERSONAL CARE & MISC": [
        "NON-ESSENTIAL – Personal Care",
        "NON-ESSENTIAL – Pet Supplies",
        "NON-ESSENTIAL – Shipping",
        "NON-ESSENTIAL – Transport",
        "NON-ESSENTIAL – Education (Language/Learning App)",
        "NON-ESSENTIAL – Education (Kids Sports)",
        "NON-ESSENTIAL – Education (Licensing Exams)",
        "NON-ESSENTIAL – Education (Online Courses)",
        "NON-ESSENTIAL – Int'l Calling (IDT)",
        "NON-ESSENTIAL – Int'l Money Transfer",
        "NON-ESSENTIAL – PayPal (misc)",
        "NON-ESSENTIAL – Other/Uncategorized",
    ],
}


def _fl(h): return PatternFill("solid", fgColor=h)
def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _hdr(ws, r, c, txt, bg="1F3864", fc="FFFFFF", sz=10, align="center"):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font = Font(bold=True, size=sz, color=fc, name="Calibri")
    cell.fill = _fl(bg); cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _bdr(); return cell
def _lbl(ws, r, c, txt, bg="FFFFFF", bold=False, fc="000000", align="left"):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font = Font(bold=bold, size=10, color=fc, name="Calibri")
    cell.fill = _fl(bg); cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _bdr(); return cell
def _mny(ws, r, c, v, bg="FFFFFF", bold=False, fc=None):
    if fc is None:
        fc = "C00000" if v < 0 else "276221" if v > 0 else "444444"
    cell = ws.cell(row=r, column=c, value=v)
    cell.number_format = "$#,##0.00"; cell.font = Font(bold=bold, size=10, color=fc, name="Calibri")
    cell.fill = _fl(bg); cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = _bdr(); return cell


def build_non_essential_report(
    all_txns: List[Dict],
    years: List[int],
    output_path: str,
):
    """
    Build the non-essential spending analysis workbook.

    Args:
        all_txns:    All transactions with 'year', 'month', 'date',
                     'desc', 'amount', 'category' keys
        years:       List of years to include (e.g. [2021,2022,2023])
        output_path: Where to write the .xlsx file
    """
    non_ess = [t for t in all_txns
               if t["category"].startswith("NON-ESSENTIAL") and t["amount"] < 0]

    by_cat = defaultdict(lambda: defaultdict(float))
    for t in non_ess:
        by_cat[t["category"]][t["year"]] += t["amount"]

    income_yr    = defaultdict(float)
    essential_yr = defaultdict(float)
    noness_yr    = defaultdict(float)
    transfer_yr  = defaultdict(float)
    for t in all_txns:
        yr = t["year"]; amt = t["amount"]; cat = t["category"]
        if cat == "INCOME" or cat.startswith("Income"):
            income_yr[yr] += amt
        elif cat.startswith("Essential") or cat.startswith("Business"):
            essential_yr[yr] += amt
        elif cat.startswith("NON-ESSENTIAL"):
            noness_yr[yr] += amt
        else:
            transfer_yr[yr] += amt

    wb = openpyxl.Workbook()
    ncols = len(years)

    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    for i in range(ncols): ws.column_dimensions[get_column_letter(i+2)].width = 14
    ws.column_dimensions[get_column_letter(ncols+2)].width = 16

    ws.merge_cells(f"A1:{get_column_letter(ncols+2)}1")
    ws["A1"].value = f"SALOME DANIEL  ·  NON-ESSENTIAL SPENDING ANALYSIS  ·  {years[0]}–{years[-1]}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = _fl("1F3864"); ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{get_column_letter(ncols+2)}2")
    ws["A2"].value = "Non-essential = spending that could have been reduced and redirected to savings  |  Powered by TAXY-Z"
    ws["A2"].font = Font(italic=True, size=9, color="333333", name="Calibri")
    ws["A2"].fill = _fl("D6E4F0"); ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    _hdr(ws, 3, 1, "Category", align="left")
    for i, yr in enumerate(years): _hdr(ws, 3, i+2, str(yr))
    _hdr(ws, 3, ncols+2, f"{len(years)}-YR TOTAL")
    ws.row_dimensions[3].height = 22

    row = 4

    def _sec(label, bg):
        nonlocal row
        ws.merge_cells(f"A{row}:{get_column_letter(ncols+2)}{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        c.fill = _fl(bg); c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bdr(); ws.row_dimensions[row].height = 18; row += 1

    # P&L overview
    _sec("▶  PROFIT & LOSS OVERVIEW", "1F3864")
    for label, data, fc, bg in [
        ("Total Income",          income_yr,    "276221", "E2EFDA"),
        ("Essential Expenses",    essential_yr, "C00000", "FCE4D6"),
        ("Non-Essential Expenses",noness_yr,   "B7472A", "FFF2CC"),
        ("Transfers / Other",     transfer_yr,  "555555", "F2F2F2"),
    ]:
        _lbl(ws, row, 1, f"  {label}", bg=bg, bold=True)
        total = 0
        for i, yr in enumerate(years):
            v = data.get(yr, 0); total += v
            _mny(ws, row, i+2, v, bg=bg, fc=fc)
        _mny(ws, row, ncols+2, total, bg=bg, bold=True, fc=fc)
        ws.row_dimensions[row].height = 16; row += 1

    # Net row
    _lbl(ws, row, 1, "  NET INCOME / (LOSS)", bg="1F3864", bold=True, fc="FFFFFF")
    gt = 0
    for i, yr in enumerate(years):
        net = income_yr.get(yr,0)+essential_yr.get(yr,0)+noness_yr.get(yr,0)+transfer_yr.get(yr,0)
        gt += net; bg = "1E6B2F" if net >= 0 else "8B0000"
        _mny(ws, row, i+2, net, bg=bg, bold=True, fc="FFFFFF")
    bg5 = "1E6B2F" if gt >= 0 else "8B0000"
    _mny(ws, row, ncols+2, gt, bg=bg5, bold=True, fc="FFFFFF")
    ws.row_dimensions[row].height = 20; row += 2

    # Potential savings
    _sec("▶  POTENTIAL SAVINGS IF NON-ESSENTIAL SPENDING ELIMINATED", "B7472A")
    _lbl(ws, row, 1, "  Non-Essential → could be savings", bg="FFF2CC", bold=True)
    total_ne = 0
    for i, yr in enumerate(years):
        v = abs(noness_yr.get(yr, 0)); total_ne += v
        _mny(ws, row, i+2, v, bg="FFF2CC", fc="B7472A", bold=True)
    _mny(ws, row, ncols+2, total_ne, bg="FFF2CC", bold=True, fc="B7472A")
    ws.row_dimensions[row].height = 16; row += 1

    _lbl(ws, row, 1, "  Monthly avg that could go to savings", bg="FEF9E7")
    for i, yr in enumerate(years):
        _mny(ws, row, i+2, abs(noness_yr.get(yr,0))/12, bg="FEF9E7", fc="B7472A")
    _mny(ws, row, ncols+2, total_ne/(12*len(years)), bg="FEF9E7", bold=True, fc="B7472A")
    ws.row_dimensions[row].height = 15; row += 2

    # Non-essential breakdown by group
    _sec("▶  NON-ESSENTIAL BREAKDOWN BY CATEGORY", "B7472A")

    for group, cats in CATEGORY_GROUPS.items():
        style = GROUP_STYLES[group]
        ws.merge_cells(f"A{row}:{get_column_letter(ncols+2)}{row}")
        c = ws.cell(row=row, column=1, value=f"  {group}")
        c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        c.fill = _fl(style["header"]); c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bdr(); ws.row_dimensions[row].height = 16; row += 1

        grp_totals = [0.0] * ncols
        for cat in cats:
            if cat not in by_cat: continue
            short = cat.replace("NON-ESSENTIAL – ", "  ")
            _lbl(ws, row, 1, short, bg=style["bg"])
            total = 0
            for i, yr in enumerate(years):
                v = by_cat[cat].get(yr, 0); total += v; grp_totals[i] += v
                if v:
                    _mny(ws, row, i+2, v, bg=style["bg"], fc=style["header"])
                else:
                    c2 = ws.cell(row=row, column=i+2, value="—")
                    c2.font = Font(color="CCCCCC", size=9, name="Calibri")
                    c2.fill = _fl(style["bg"]); c2.alignment = Alignment(horizontal="center")
                    c2.border = _bdr()
            _mny(ws, row, ncols+2, total, bg=style["bg"], bold=True, fc=style["header"])
            ws.row_dimensions[row].height = 14; row += 1

        _lbl(ws, row, 1, f"  {group} SUBTOTAL", bg=style["header"], bold=True, fc="FFFFFF")
        gt2 = 0
        for i, yr in enumerate(years):
            _mny(ws, row, i+2, grp_totals[i], bg=style["header"], bold=True, fc="FFFFFF"); gt2 += grp_totals[i]
        _mny(ws, row, ncols+2, gt2, bg=style["header"], bold=True, fc="FFFFFF")
        ws.row_dimensions[row].height = 16; row += 1

    # Grand total
    row += 1
    _lbl(ws, row, 1, "TOTAL NON-ESSENTIAL SPENDING", bg="1F3864", bold=True, fc="FFFFFF", align="center")
    grand = 0
    for i, yr in enumerate(years):
        v = noness_yr.get(yr, 0); grand += v
        _mny(ws, row, i+2, v, bg="1F3864", bold=True, fc="FFFFFF")
    _mny(ws, row, ncols+2, grand, bg="1F3864", bold=True, fc="FFFFFF")
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = "B4"

    # ── All Non-Essential Transactions sheet ──────────────────────────────────
    ws2 = wb.create_sheet("All Non-Essential Txns")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [6,18,16,56,36,12]): ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:F1")
    ws2["A1"].value = f"ALL NON-ESSENTIAL TRANSACTIONS — {years[0]}–{years[-1]} — {len(non_ess)} items"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    ws2["A1"].fill = _fl("1F3864"); ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for col, txt in enumerate(["Year","Month","Date","Description","Category","Amount"], 1):
        _hdr(ws2, 2, col, txt)
    ws2.row_dimensions[2].height = 18

    def _row_bg(cat):
        c = cat.upper()
        if "FOOD" in c or "RESTAURANT" in c: return "FCE4D6"
        if "SHOPPING" in c: return "FFF2CC"
        if "SUBSCRIPTION" in c: return "DAE3F3"
        if "TRAVEL" in c: return "E2EFDA"
        if "ENTERTAINMENT" in c: return "EAD1DC"
        return "F2F2F2"

    tr = 3
    for t in sorted(non_ess, key=lambda x: (x["year"], x["month"], x["date"])):
        bg = _row_bg(t["category"])
        _lbl(ws2, tr, 1, str(t["year"]), bg=bg, align="center")
        _lbl(ws2, tr, 2, t["month"].replace(f" {t['year']}", ""), bg=bg)
        _lbl(ws2, tr, 3, t["date"], bg=bg, align="center")
        _lbl(ws2, tr, 4, t["desc"], bg=bg)
        _lbl(ws2, tr, 5, t["category"].replace("NON-ESSENTIAL – ", ""), bg=bg)
        _mny(ws2, tr, 6, t["amount"], bg=bg)
        ws2.row_dimensions[tr].height = 13; tr += 1

    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:F{tr-1}"

    wb.save(output_path)
    print(f"✅ Non-essential analysis saved: {output_path}  ({len(non_ess)} non-essential transactions)")

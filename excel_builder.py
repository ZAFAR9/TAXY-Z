"""
TAXY-Z — Excel Report Builder
================================
Generates styled, professional Excel P&L workbooks from categorized
transaction data. Produces:
  - Annual Summary sheet (month × category matrix)
  - All Transactions sheet (filterable full list)
  - Individual monthly P&L sheets

Usage:
    from excel_builder import build_annual_report
    build_annual_report(all_months_data, year=2024, output_path="PL_2024.xlsx")
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from typing import List, Dict


# ── Palette ──────────────────────────────────────────────────────────────────
PALETTE = {
    "navy":       "1F3864",
    "dark_green": "1E6B2F",
    "dark_red":   "8B0000",
    "mid_blue":   "2E75B6",
    "light_blue": "D6E4F0",
    "income_bg":  "E2EFDA",
    "expense_bg": "FCE4D6",
}

CAT_COLORS = {
    "Mortgage / Housing":             "B8CCE4",
    "Essential – Loan Payment":       "FFE699",
    "Essential – Rent Paid":          "BDD7EE",
    "Essential – Groceries":          "E9F7EF",
    "Essential – Utilities":          "E8D5B7",
    "Essential – Utilities (Phone/Internet)": "E8D5B7",
    "Essential – Insurance (Auto)":   "FFC7CE",
    "Essential – Insurance (Life)":   "F4CCCC",
    "Essential – Insurance (Home)":   "FFC7CE",
    "Essential – Insurance (Pet)":    "F4CCCC",
    "Essential – Insurance (Other)":  "FFC7CE",
    "Essential – Home Security":      "FFC7CE",
    "Essential – Health (Kaiser)":    "FADBD8",
    "Essential – Health (Dental)":    "FADBD8",
    "Essential – Health (Vet)":       "FADBD8",
    "Essential – Pharmacy":           "FADBD8",
    "Essential – Auto (Gas)":         "FFE0B2",
    "Essential – Auto (DMV)":         "FFE0B2",
    "Essential – Auto Repair":        "FFE0B2",
    "Essential – Education (School)": "C6EFCE",
    "Essential – Credit Card Payment":"E0E0E0",
    "Essential – Bank Fees":          "BDBDBD",
    "Essential – Taxes":              "F1948A",
    "Essential – Legal":              "F1948A",
    "Essential – Home Maintenance":   "E8D5B7",
    "Essential – Equipment Lease":    "E8D5B7",
    "NON-ESSENTIAL – Food Delivery (DoorDash)":     "FCF3CF",
    "NON-ESSENTIAL – Restaurants & Dining Out":     "FCF3CF",
    "NON-ESSENTIAL – Shopping (Amazon)":            "F9EBEA",
    "NON-ESSENTIAL – Shopping (Clothing)":          "F9EBEA",
    "NON-ESSENTIAL – Shopping (Jewelry)":           "F9EBEA",
    "NON-ESSENTIAL – Shopping (Electronics)":       "F9EBEA",
    "NON-ESSENTIAL – Shopping (Beauty)":            "F9EBEA",
    "NON-ESSENTIAL – Shopping (Walmart)":           "F9EBEA",
    "NON-ESSENTIAL – Subscriptions (Amazon Prime)": "D4E6F1",
    "NON-ESSENTIAL – Subscriptions (Streaming)":    "D4E6F1",
    "NON-ESSENTIAL – Subscriptions (Gaming)":       "D4E6F1",
    "NON-ESSENTIAL – Travel (Flights)":             "E8DAEF",
    "NON-ESSENTIAL – Travel (Hotel)":               "E8DAEF",
    "NON-ESSENTIAL – Travel (Airbnb)":              "E8DAEF",
    "NON-ESSENTIAL – Entertainment (Events)":       "E8DAEF",
    "NON-ESSENTIAL – PayPal (misc)":                "F5CBA7",
    "NON-ESSENTIAL – Int'l Calling (IDT)":          "D7BDE2",
    "NON-ESSENTIAL – Int'l Money Transfer":         "D7BDE2",
    "Business – Merchant Fees":                     "D5D8DC",
    "Savings Transfer":                             "AED6F1",
    "Internal Transfer":                            "F0F0F0",
    "Transfer – Zelle Out":                         "FDEBD0",
    "Transfer – Domestic Wire Out":                 "F1948A",
    "Cash Withdrawal":                              "E0E0E0",
    "Check Written":                                "E0E0E0",
}


# ── Style helpers ─────────────────────────────────────────────────────────────
def _fl(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _bdr(color: str = "CCCCCC") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, col, text, bg=PALETTE["navy"], fc="FFFFFF", sz=10,
         bold=True, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, size=sz, color=fc, name="Calibri")
    c.fill = _fl(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _bdr()
    return c


def _lbl(ws, row, col, text, bg="FFFFFF", bold=False, fc="000000", align="left"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, size=10, color=fc, name="Calibri")
    c.fill = _fl(bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _bdr()
    return c


def _mny(ws, row, col, value, bg="FFFFFF", bold=False, fc=None):
    if fc is None:
        fc = "C00000" if value < 0 else "276221" if value > 0 else "666666"
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = "$#,##0.00"
    c.font = Font(bold=bold, size=10, color=fc, name="Calibri")
    c.fill = _fl(bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _bdr()
    return c


# ── Main builder ──────────────────────────────────────────────────────────────
def build_annual_report(
    all_months: List[Dict],
    year: int,
    output_path: str,
    income_cats: List[str],
    expense_cats: List[str],
):
    """
    Build a full-year P&L Excel workbook.

    Args:
        all_months:   List of month dicts: {name, period, txns}
                      Each txn: {date, desc, amount, category}
        year:         Calendar year (e.g. 2024)
        output_path:  Where to save the .xlsx file
        income_cats:  Ordered list of income category strings
        expense_cats: Ordered list of expense category strings
    """
    wb = openpyxl.Workbook()
    ncols = len(all_months)
    month_names = [m["name"].replace(f" {year}", "") for m in all_months]
    txn_total = sum(len(m["txns"]) for m in all_months)

    # Per-month category totals
    month_cat_totals = []
    for md in all_months:
        ct = defaultdict(float)
        for t in md["txns"]:
            ct[t["category"]] += t["amount"]
        month_cat_totals.append(ct)

    # ── Annual Summary ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Annual Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(ncols + 2)}1")
    ws["A1"].value = f"SALOME DANIEL  ·  PROFIT & LOSS  ·  {year}  ·  Chase Account #0828"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = _fl(PALETTE["navy"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(ncols + 2)}2")
    ws["A2"].value = (
        f"All {txn_total} transactions from 12 Chase statements — "
        "parsed & categorized from PDF source data  |  Powered by TAXY-Z"
    )
    ws["A2"].font = Font(italic=True, size=9, color="333333", name="Calibri")
    ws["A2"].fill = _fl(PALETTE["light_blue"])
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    _hdr(ws, 4, 1, "Category", align="left")
    for i, mn in enumerate(month_names):
        _hdr(ws, 4, i + 2, mn, sz=9)
    _hdr(ws, 4, ncols + 2, "ANNUAL TOTAL")
    ws.row_dimensions[4].height = 26
    ws.column_dimensions["A"].width = 38
    for i in range(ncols + 1):
        ws.column_dimensions[get_column_letter(i + 2)].width = 11

    row = 5

    def _section(label, bg):
        nonlocal row
        ws.merge_cells(f"A{row}:{get_column_letter(ncols + 2)}{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        c.fill = _fl(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    # Income section
    _section("▶  INCOME", PALETTE["dark_green"])
    total_income_pm = [0.0] * ncols
    for cat in income_cats:
        if not any(month_cat_totals[i].get(cat, 0) for i in range(ncols)):
            continue
        _lbl(ws, row, 1, f"  {cat}", bg=PALETTE["income_bg"])
        ann = 0.0
        for i in range(ncols):
            v = month_cat_totals[i].get(cat, 0.0)
            ann += v
            total_income_pm[i] += v
            _mny(ws, row, i + 2, v, bg=PALETTE["income_bg"],
                 fc="276221" if v > 0 else "666666")
        _mny(ws, row, ncols + 2, ann, bg=PALETTE["income_bg"],
             bold=True, fc="276221" if ann > 0 else "666666")
        ws.row_dimensions[row].height = 15
        row += 1

    _lbl(ws, row, 1, "TOTAL INCOME", bg=PALETTE["dark_green"], bold=True,
         fc="FFFFFF", align="center")
    ann_inc = sum(total_income_pm)
    for i in range(ncols):
        _mny(ws, row, i + 2, total_income_pm[i], bg=PALETTE["dark_green"],
             bold=True, fc="FFFFFF")
    _mny(ws, row, ncols + 2, ann_inc, bg=PALETTE["dark_green"],
         bold=True, fc="FFFFFF")
    ws.row_dimensions[row].height = 20
    row += 1

    # Expense section
    _section("▶  EXPENSES", PALETTE["dark_red"])
    total_exp_pm = [0.0] * ncols
    for cat in expense_cats:
        if not any(month_cat_totals[i].get(cat, 0) for i in range(ncols)):
            continue
        bg = CAT_COLORS.get(cat, "F0F0F0")
        _lbl(ws, row, 1, f"  {cat}", bg=bg)
        ann = 0.0
        for i in range(ncols):
            v = month_cat_totals[i].get(cat, 0.0)
            ann += v
            total_exp_pm[i] += v
            if v:
                _mny(ws, row, i + 2, v, bg=bg)
            else:
                c2 = ws.cell(row=row, column=i + 2, value="—")
                c2.font = Font(color="CCCCCC", size=9, name="Calibri")
                c2.fill = _fl(bg)
                c2.alignment = Alignment(horizontal="center")
                c2.border = _bdr()
        _mny(ws, row, ncols + 2, ann, bg=bg, bold=True)
        ws.row_dimensions[row].height = 15
        row += 1

    _lbl(ws, row, 1, "TOTAL EXPENSES", bg=PALETTE["dark_red"], bold=True,
         fc="FFFFFF", align="center")
    ann_exp = sum(total_exp_pm)
    for i in range(ncols):
        _mny(ws, row, i + 2, total_exp_pm[i], bg=PALETTE["dark_red"],
             bold=True, fc="FFFFFF")
    _mny(ws, row, ncols + 2, ann_exp, bg=PALETTE["dark_red"],
         bold=True, fc="FFFFFF")
    ws.row_dimensions[row].height = 20
    row += 1

    # Net row
    row += 1
    ann_net = ann_inc + ann_exp
    net_bg = PALETTE["dark_green"] if ann_net >= 0 else PALETTE["dark_red"]
    _lbl(ws, row, 1, "NET INCOME / (LOSS)", bg=net_bg, bold=True,
         fc="FFFFFF", align="center")
    for i in range(ncols):
        net = total_income_pm[i] + total_exp_pm[i]
        bg = PALETTE["dark_green"] if net >= 0 else PALETTE["dark_red"]
        _mny(ws, row, i + 2, net, bg=bg, bold=True, fc="FFFFFF")
    _mny(ws, row, ncols + 2, ann_net, bg=net_bg, bold=True, fc="FFFFFF")
    ws.row_dimensions[row].height = 24
    ws.freeze_panes = "B5"

    # ── All Transactions sheet ────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Transactions")
    ws_all.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [14, 16, 54, 14, 38]):
        ws_all.column_dimensions[col].width = w

    ws_all.merge_cells("A1:E1")
    ws_all["A1"].value = (
        f"ALL TRANSACTIONS — SALOME DANIEL {year} ({txn_total} lines)"
    )
    ws_all["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    ws_all["A1"].fill = _fl(PALETTE["navy"])
    ws_all["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_all.row_dimensions[1].height = 26

    for col, txt in enumerate(["Month", "Date", "Description", "Amount", "Category"], 1):
        _hdr(ws_all, 2, col, txt, sz=10)
    ws_all.row_dimensions[2].height = 20

    tr = 3
    for md in all_months:
        mn = md["name"].replace(f" {year}", "")
        for t in md["txns"]:
            bg = CAT_COLORS.get(t["category"], "F9F9F9") if t["amount"] < 0 else "E8F5E9"
            _lbl(ws_all, tr, 1, mn, bg=bg)
            _lbl(ws_all, tr, 2, t["date"], bg=bg, align="center")
            _lbl(ws_all, tr, 3, t["desc"], bg=bg)
            _mny(ws_all, tr, 4, t["amount"], bg=bg)
            _lbl(ws_all, tr, 5, t["category"], bg=bg)
            ws_all.row_dimensions[tr].height = 13
            tr += 1

    ws_all.freeze_panes = "A3"
    ws_all.auto_filter.ref = f"A2:E{tr - 1}"

    # ── Monthly sheets ────────────────────────────────────────────────────────
    for md, ct in zip(all_months, month_cat_totals):
        short = md["name"].replace(f" {year}", "")
        ws_m = wb.create_sheet(short)
        ws_m.sheet_view.showGridLines = False
        for col, w in zip("ABCD", [40, 14, 8, 10]):
            ws_m.column_dimensions[col].width = w

        ws_m.merge_cells("A1:D1")
        ws_m["A1"].value = f"{md['name'].upper()} — P&L"
        ws_m["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
        ws_m["A1"].fill = _fl(PALETTE["navy"])
        ws_m["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_m.row_dimensions[1].height = 26

        ws_m.merge_cells("A2:D2")
        ws_m["A2"].value = f"Statement Period: {md['period']}"
        ws_m["A2"].font = Font(italic=True, size=9, color="333333", name="Calibri")
        ws_m["A2"].fill = _fl(PALETTE["light_blue"])
        ws_m["A2"].alignment = Alignment(horizontal="center")

        for col, txt in enumerate(["Category", "Amount", "Txns", ""], 1):
            _hdr(ws_m, 3, col, txt, sz=10)
        ws_m.row_dimensions[3].height = 18

        r = 4
        for section_label, cats, section_bg in [
            ("INCOME", income_cats, PALETTE["dark_green"]),
            ("EXPENSES", expense_cats, PALETTE["dark_red"]),
        ]:
            ws_m.merge_cells(f"A{r}:D{r}")
            ws_m[f"A{r}"].value = section_label
            ws_m[f"A{r}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            ws_m[f"A{r}"].fill = _fl(section_bg)
            ws_m[f"A{r}"].alignment = Alignment(horizontal="left")
            ws_m.row_dimensions[r].height = 18
            r += 1

            section_total = 0.0
            for cat in cats:
                v = ct.get(cat, 0.0)
                if not v:
                    continue
                n = sum(1 for t in md["txns"] if t["category"] == cat)
                bg = PALETTE["income_bg"] if section_label == "INCOME" else CAT_COLORS.get(cat, "F0F0F0")
                fc = "276221" if section_label == "INCOME" else None
                _lbl(ws_m, r, 1, f"  {cat}", bg=bg)
                _mny(ws_m, r, 2, v, bg=bg, fc=fc)
                _lbl(ws_m, r, 3, str(n), bg=bg, align="center")
                _lbl(ws_m, r, 4, "", bg=bg)
                section_total += v
                r += 1

            _lbl(ws_m, r, 1, f"TOTAL {section_label}", bg=section_bg,
                 bold=True, fc="FFFFFF", align="center")
            _mny(ws_m, r, 2, section_total, bg=section_bg, bold=True, fc="FFFFFF")
            _lbl(ws_m, r, 3, "", bg=section_bg)
            _lbl(ws_m, r, 4, "", bg=section_bg)
            ws_m.row_dimensions[r].height = 20
            r += 2

        ws_m.freeze_panes = "A4"

    wb.save(output_path)
    print(f"✅ Report saved: {output_path}  ({txn_total} transactions)")
